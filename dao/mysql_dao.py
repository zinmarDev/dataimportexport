import mysql.connector
from mysql.connector import Error
from xlsx2csv import Xlsx2csv
import pandas as pd
import csv
from openpyxl import load_workbook


def get_connection(db_config):
    host = db_config.get("Host")
    port = db_config.get("Port")
    user_name = db_config.get("UserName")
    password = db_config.get("Password")
    database = db_config.get("DatabaseName")
    print("host : ", host, " and port : ", port)
    error = ""
    try:
        connection = mysql.connector.connect(host=host,
                                             port=port,
                                             database=database,
                                             user=user_name,
                                             password=password)

        if connection.is_connected():
            db_info = connection.get_server_info()
            print("Connected to MySQL Server version ", db_info)
            cursor = connection.cursor()
            cursor.execute("select database();")
            record = cursor.fetchone()
            print("You're connected to database: ", record)

            return connection, error
        else:
            return connection, error
    except Error as e:
        print("Error while connecting to MySQL", e)
        error = e
        return None, error


def create_table(db_config):
    connection, err = get_connection(db_config)
    cursor = connection.cursor()
    err = ""
    try:
        table_info = db_config.get("TableInfo")  # table info may be array.
        create_table_query = ""
        for ind in range(len(table_info)):
            info = table_info[ind]
            create_table_query = "CREATE TABLE " + info.get("TableName")
            column_info = ""
            unique_key = ""
            primary_key = ""
            foreign_key = ""
            for col_info in info.get("Columns"):
                if column_info != "":
                    column_info += ","

                if col_info.get("PKey"):
                    column_info += col_info.get("ColumnName") + " " + col_info.get("DataType") + " NOT NULL"

                elif col_info.get("UKey"):
                    column_info += col_info.get("ColumnName") + " " + col_info.get("DataType") + " NOT NULL"

                else:
                    column_info += col_info.get("ColumnName") + " " + col_info.get("DataType")

                if col_info.get("PKey"):
                    primary_key = get_primary_key(col_info, primary_key)

                if col_info.get("UKey"):
                    unique_key = get_unique_key(col_info, unique_key)

                if col_info.get("FKey"):
                    foreign_key = get_foreign_key(col_info, foreign_key)

                if primary_key != "":
                    column_info += ", PRIMARY KEY(" + primary_key + ")"

                if unique_key != "":
                    column_info += ", UNIQUE (" + unique_key + ")"

                if foreign_key != "":
                    column_info += ", " + foreign_key

            create_table_query += "(" + column_info + ")"
        print("create table query : ", create_table_query)

        result = cursor.execute(create_table_query)
        print("Laptop Table created successfully ", result)

    except mysql.connector.Error as error:
        err = str(error)
        print("Failed to create table in MySQL: {}".format(error))
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
            print("MySQL connection is closed")
    return err


def get_foreign_key(col_info, f_key):
    foreign_key = ""
    if col_info.get("FKey"):
        foreign_key += "FOREIGN KEY(" + col_info.get("ColumnName") + ") REFERENCES " \
                       + col_info.get("RefTable") + "(" + col_info.get("RefColumn") + ")"

    if foreign_key != "" and f_key != "":
        foreign_key += ", " + foreign_key

    return foreign_key


def get_primary_key(col_info, primary_key):
    if primary_key != "":
        primary_key += "," + col_info.get("ColumnName")
    else:
        primary_key += col_info.get("ColumnName")
    return primary_key


def get_unique_key(col_info, unique_key):
    if unique_key != "":
        unique_key += "," + col_info.get("ColumnName")
    else:
        unique_key += col_info.get("ColumnName")
    return unique_key


def prepare_conditions(existed_cond, conditions, join_operator):
    condition = existed_cond
    if conditions != "" and conditions is not None and len(conditions) > 0:
        for cond in conditions:
            if condition == "":
                condition = cond.get("ColumnName") + cond.get("Operator") + "'" + str(cond.get("Value")) + "'"
            else:
                condition += " " + join_operator + " " + cond.get("ColumnName") + cond.get("Operator") + \
                             "'" + cond.get("Value") + "'"
    return condition


def export_data_file(connection, config):
    columns = "*"
    conditions = ""
    conditions = prepare_conditions(conditions, config.get("ANDConditions"), "AND")
    conditions = prepare_conditions(conditions, config.get("ORConditions"), "OR")

    if config.get("ColumnList") is not None and len(config.get("ColumnList")) > 0:
        columns = ""
        for col in config.get("ColumnList"):
            if columns != "":
                columns += ", "+col
            else:
                columns += col

    query = "SELECT " + columns + " FROM " + config.get("TableName")
    if conditions != "":
        query += " WHERE " + conditions

    query += " INTO OUTFILE '"+config.get("FilePath") + "' FIELDS TERMINATED BY '"+ config.get("Delimiter") + "'"
    query += " LINES TERMINATED BY '\r\n'"

    print("export data query : ", query)

    cursor = connection.cursor()
    err = ""
    try:
        result = cursor.execute(query)
        print("query is executed successfully ", result)

    except mysql.connector.Error as error:
        err = error
        print("Failed to create table in MySQL: {}".format(error))
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
            print("MySQL connection is closed")
    return "data are exported successfully", err


def import_data_file(connection, config):
    query = "LOAD DATA INFILE '" + config.get("FilePath") + "' INTO TABLE "+ config.get("TableName") + " FIELDS TERMINATED BY '" + config.get("Delimiter") + "'"

    if config.get("RemoveFirstRow"):
        query += " IGNORE 1 LINES"
    print("import data query : ", query)

    cursor = connection.cursor()
    err = ""
    try:
        cursor.execute("use "+config.get("DatabaseName"))
        cursor.execute(query)
        connection.commit()
        print("query is executed successfully ")

    except mysql.connector.Error as error:
        err = error
        print("Failed to create table in MySQL: {}".format(error))
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
            print("MySQL connection is closed")
    return "data is imported successfully", err


def convert_xlsx_csv(config):
    from_files = config.get("FromFilePath")
    to_files = config.get("ToFilePath")
    if len(from_files) == len(to_files):
        for ind in range(len(from_files)):
            from_file = from_files[ind]
            to_file = to_files[ind]
            Xlsx2csv(from_file).convert(to_file)
        return "converted xlsx to csv file successfully.", ""
    else:
        return "converted xlsx to csv file successfully.", "must be equal from/to file paths"


def convert_xlsx_csv_one(config):
    from_files = config.get("FromFilePath")
    to_files = config.get("ToFilePath")

    for ind in range(len(from_files)):
        wb = load_workbook(filename=from_files[ind])
        sheet = wb.active
        csv_data = []
        for value in sheet.iter_rows(values_only=True):
            csv_data.append(list(value))

        with open(to_files[ind], "w") as csv_obj:
            writer = csv.writer(csv_obj)
            for line in csv_data:
                writer.writerow(line)
    return "converted xlsx to csv file successfully.", ""


def convert_sql_file(config):
    from_files = config.get("FromFilePath")
    to_files = config.get("ToFilePath")

    for ind in range(len(from_files)):
        from_file = from_files[ind]
        to_file = to_files[ind]
        data = pd.read_excel(from_file)
        print("columns : ", data.columns, "\n heads : ", data.head(), "n data : ", data.describe())
        # df = pd.DataFrame(data)
        # print(df)

    return "converted xlsx to sql file successfully.", ""



